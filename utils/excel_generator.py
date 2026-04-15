import openpyxl
import math
from collections import defaultdict
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from utils.draw_bracket import get_seeded_participants
import html
from .formatters import format_weight, format_fio_without_patronymic
from datetime import datetime

RANK_ABBREVIATIONS = {
    "III юношеский": "3юн",
    "II юношеский": "2юн",
    "I юношеский": "1юн",
    "III спортивный": "3сп",
    "II спортивный": "2сп",
    "I спортивный": "1сп",
    "КМС": "КМС",
    "МС": "МС",
    "МСМК": "МСМК",
    "ЗМС": "ЗМС",
}

def _rank_suffix(rank_title: str) -> str:
    """Возвращает сокращение разряда, если известно, иначе исходное название.
    Возвращает пустую строку, если вход пустой/None.
    """
    if not rank_title:
        return ""
    title = str(rank_title).strip()
    # Пробуем точное совпадение
    if title in RANK_ABBREVIATIONS:
        return RANK_ABBREVIATIONS[title]
    # Пробуем нормализацию регистра
    title_cap = title.capitalize()
    if title_cap in RANK_ABBREVIATIONS:
        return RANK_ABBREVIATIONS[title_cap]
    # Пробуем частые варианты записи и нормализацию латинской І -> I
    title_normalized = title.replace("І", "I").replace("і", "I")
    alt = (
        title_normalized.replace(" юн.", " юношеский").replace(" юн", " юношеский")
             .replace(" спорт.", " спортивный").replace(" спорт", " спортивный")
             .replace("III", "III").replace("II", "II").replace("I", "I")
    )
    if alt in RANK_ABBREVIATIONS:
        return RANK_ABBREVIATIONS[alt]
    # Если не нашли в словаре, вернем исходное название (может быть в другой записи)
    return title

# --- Помощники для сортировки ---
def _age_sort_key(age_str: str):
    if not age_str: return float('-inf')
    if age_str == "Без категории":
        return float('-inf')
    if 'старше' in age_str:
        try:
            return int(age_str.split(' ')[1])
        except (ValueError, IndexError):
            return float('-inf')
    try:
        return int(age_str.split('-')[0])
    except (ValueError, IndexError):
        return float('-inf')

def _weight_sort_key(weight_str: str | None):
    if not weight_str: return float('inf')
    try:
        return float(weight_str.replace('+', '.1'))
    except ValueError:
        return float('inf')

def _class_sort_key(class_str: str):
    """Возвращает числовой ключ для сортировки классов (А -> В -> С)."""
    if not class_str: return 99 # Обработка пустого значения
    # Исправлен порядок: 'А' (0) -> 'В' (1) -> 'С' (2) для стандартной сортировки
    order = {'А': 0, 'В': 1, 'С': 2}
    return order.get(class_str.split(' ')[0], 99)


def _get_stage_text(num_rounds: int, current_round: int) -> str:
    """Возвращает текстовое представление стадии турнира (Финал, 1/2, 1/4...)."""
    if current_round >= num_rounds:
        return "\nФинал"

    num_fights_in_round = 2 ** (num_rounds - current_round)
    if num_fights_in_round > 1:
        return f"\n1/{num_fights_in_round}"

    return "\nФинал"

def _get_stage_sort_key(num_rounds: int, current_round: int) -> int:
    """Возвращает числовой ключ для сортировки по стадиям (1/8 -> 1/4 -> 1/2 -> Финал)."""
    fights_in_round = 2 ** (num_rounds - current_round)
    if fights_in_round >= 8: # 1/8
        return 0
    elif fights_in_round == 4: # 1/4
        return 1
    elif fights_in_round == 2: # 1/2
        return 2
    else: # Финал
        return 3

def generate_preliminary_list_excel(participants: list, file_path: str):
    """
    Генерирует Excel-файл с предварительным списком участников,
    разделенный по возрастным категориям и полу, с автоподбором
    ширины столбцов.
    """
    # 1. Группировка участников по возрастной категории
    grouped_by_age = defaultdict(list)
    for p in participants:
        # Участники без категории попадают в лист "Без категории"
        age_category = p.get('age_category_name') or "Без категории"
        grouped_by_age[age_category].append(p)

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # Удаляем стандартный лист

    # Сортируем категории так, чтобы лист "8-9" был первым
    sorted_age_categories = sorted(list(grouped_by_age.keys()))
    if "8-9" in sorted_age_categories:
        sorted_age_categories.remove("8-9")
        sorted_age_categories.insert(0, "8-9")

    # 2. Создание листов для каждой возрастной категории
    for age_category in sorted_age_categories:
        sheet = workbook.create_sheet(title=age_category)

        # Заголовки таблицы с учетом столбца "Категория"
        headers = ["ФИО", "Дата рождения", "Вес", "Категория", "Пол", "Клуб", "Город"]
        sheet.append(headers)

        # Стилизация заголовков (жирный шрифт, выравнивание по центру)
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

            # 3. Разделение участников на муж и жен
        participants_in_category = grouped_by_age[age_category]
        men = [p for p in participants_in_category if p['gender'] == 'Мужской']
        women = [p for p in participants_in_category if p['gender'] == 'Женский']

        row_num = 2  # Начинаем запись данных со второй строки

        # Добавление муж в таблицу
        if men:
            sheet.cell(row=row_num, column=1, value='Муж').font = Font(bold=True)  # Заменено
            row_num += 1
            for p in men:
                sheet.append([
                    p.get('fio'),
                    p.get('dob').strftime('%d.%m.%Y') if p.get('dob') else '',
                    p.get('weight'),
                    p.get('class_name'),  # Возвращенный столбец
                    p.get('gender'),
                    p.get('club_name'),
                    p.get('city_name')
                ])
                row_num += 1

        # Добавление женщин в таблицу
        if women:
            sheet.cell(row=row_num, column=1, value='Жен').font = Font(bold=True)
            row_num += 1
            for p in women:
                sheet.append([
                    p.get('fio'),
                    p.get('dob').strftime('%d.%m.%Y') if p.get('dob') else '',
                    p.get('weight'),
                    p.get('class_name'),  # Возвращенный столбец
                    p.get('gender'),
                    p.get('club_name'),
                    p.get('city_name')
                ])
                row_num += 1

        # 4. Автоматическая настройка ширины столбцов
        for column_cells in sheet.columns:
            # Находим максимальную длину значения в столбце
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            # Устанавливаем ширину с небольшим запасом
            column_letter = get_column_letter(column_cells[0].column)
            sheet.column_dimensions[column_letter].width = max_length + 2

    # 5. Сохранение файла
    workbook.save(file_path)

def generate_weigh_in_list_excel(participants: list, file_path: str):
    grouped_by_club = defaultdict(list)
    for p in participants:
        club_name = p.get('club_name') or "Без клуба"
        grouped_by_club[club_name].append(p)

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    sorted_clubs = sorted(list(grouped_by_club.keys()))

    for club_name in sorted_clubs:
        participants_in_club = sorted(grouped_by_club[club_name], key=lambda x: x.get('fio', ''))
        if not participants_in_club:
             continue

        safe_club_name = "".join(c for c in club_name if c.isalnum() or c in ' _-').strip()[:31]
        sheet = workbook.create_sheet(title=safe_club_name)

        headers = ["ФИО", "Дата рождения", "Вес", "Категория", "Пол", "Клуб", "Город"]
        sheet.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for p in participants_in_club:
            sheet.append([
                p.get('fio'),
                p.get('dob').strftime('%d.%m.%Y') if p.get('dob') else '',
                format_weight(p.get('weight')), # Используем format_weight
                p.get('class_name'),
                p.get('gender'),
                p.get('club_name'),
                p.get('city_name')
            ])

        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells if cell.value is not None)
            column_letter = get_column_letter(column_cells[0].column)
            sheet.column_dimensions[column_letter].width = max_length + 2

    workbook.save(file_path)


def generate_all_brackets_excel(grouped_by_category: dict, file_path: str):
    """
    Генерирует Excel-файл со всеми турнирными сетками с динамическим расчетом высоты.
    Каждая возрастная категория выносится на отдельный лист.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Удаляем лист, созданный по умолчанию

    # --- Константы для стилей и компоновки ---
    HEADER_FONT = Font(name='Calibri', size=11, bold=True)
    PARTICIPANT_FONT = Font(name='Calibri', size=11)
    SMALL_FONT = Font(name='Calibri', size=10) # Для клуба/города
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
    THIN_SIDE = Side(style='thin', color="000000")
    BORDER_BOTTOM = Border(bottom=THIN_SIDE)
    BORDER_TOP_BOTTOM = Border(top=THIN_SIDE, bottom=THIN_SIDE)
    BORDER_LEFT = Border(left=THIN_SIDE)
    BORDER_LEFT_BOTTOM = Border(left=THIN_SIDE, bottom=THIN_SIDE)

    GRID_PADDING_ROWS = 5         # Отступ между сетками (увеличено)
    CATEGORY_HEADER_ROWS = 2      # Высота заголовка категории
    SIGNATURE_BLOCK_ROWS = 2      # Количество строк для подписей
    SIGNATURE_PADDING_ROWS = 2    # Отступ перед подписями
    PARTICIPANT_ROW_HEIGHT = 18   # Высота строки для имени/клуба участника
    LINE_ROW_HEIGHT = 15          # Стандартная высота строки для линий

    # --- Сначала группируем все категории по возрасту ---
    brackets_by_age = defaultdict(list)
    for category_key, participants in grouped_by_category.items():
        # Проверяем корректность ключа
        if len(category_key) < 3 or not category_key[2]:
             print(f"Предупреждение: Некорректный ключ категории пропущен: {category_key}")
             continue
        age_cat_name = category_key[2]
        brackets_by_age[age_cat_name].append((category_key, participants))

    # Сортируем возрастные категории для правильного порядка вкладок
    sorted_age_categories = sorted(brackets_by_age.keys(), key=_age_sort_key)

    # --- Основной цикл: одна итерация = одна вкладка (возрастная категория) ---
    for age_category in sorted_age_categories:
        categories_on_sheet = brackets_by_age[age_category]
        if not categories_on_sheet:
            continue # Пропускаем создание листа, если нет категорий

        # Сортируем категории внутри листа: Пол -> Вес -> Категория
        categories_on_sheet.sort(key=lambda item: (
            0 if item[0][1] == 'Женский' else 1, # Сначала женщины
            _weight_sort_key(item[0][3]),        # Затем по весу
            _class_sort_key(item[0][0])          # Затем по классу
        ))

        # Создаем лист с безопасным именем (Excel не любит некоторые символы)
        safe_sheet_title = "".join(c for c in age_category if c.isalnum() or c in ' _-').strip()[:31]
        sheet = wb.create_sheet(title=safe_sheet_title)
        sheet.default_row_height = LINE_ROW_HEIGHT # Стандартная высота для большинства строк

        # --- Настройка колонок под шаблон ---
        sheet.column_dimensions[get_column_letter(1)].width = 35 # Имя участника 1
        sheet.column_dimensions[get_column_letter(2)].width = 3  # Коннектор 1
        sheet.column_dimensions[get_column_letter(3)].width = 35 # Имя участника 2 / Победитель R1
        sheet.column_dimensions[get_column_letter(4)].width = 3  # Коннектор 2
        sheet.column_dimensions[get_column_letter(5)].width = 35 # Победитель R2
        sheet.column_dimensions[get_column_letter(6)].width = 3  # Коннектор 3
        sheet.column_dimensions[get_column_letter(7)].width = 35 # Победитель R3
        sheet.column_dimensions[get_column_letter(8)].width = 3  # Коннектор 4
        sheet.column_dimensions[get_column_letter(9)].width = 35 # Победитель R4 (Финалист)
        # Добавляем колонки для подписей справа, если нужно больше места
        sheet.column_dimensions[get_column_letter(11)].width = 12 # "Гл. судья"
        sheet.column_dimensions[get_column_letter(12)].width = 40 # Линия и ФИО судьи

        current_row = 2 # Начинаем со второй строки для общего заголовка

        # --- Общий заголовок для листа ---
        sheet.merge_cells(f'A{current_row}:L{current_row}') # Расширяем до последней колонки
        cell = sheet.cell(current_row, 1, 'Чемпионат и Первенство республики Башкортостан по муайтай')
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        current_row += 1
        sheet.merge_cells(f'B{current_row}:K{current_row}')
        cell = sheet.cell(current_row, 2, '31.10 - 03.11.2025')
        cell.font = PARTICIPANT_FONT
        cell.alignment = Alignment(horizontal='right', vertical='center')
        current_row += 1
        sheet.merge_cells(f'A{current_row}:L{current_row}')
        cell = sheet.cell(current_row, 1, 'номер-код вида спорта: 0170001511Я')
        cell.font = PARTICIPANT_FONT
        cell.alignment = CENTER_ALIGN
        current_row += 2 # Отступ перед первой сеткой

        # --- Внутренний цикл: отрисовка всех сеток для текущей возрастной категории ---
        for category_key, participants in categories_on_sheet:
            if not participants: continue

            # Убираем None из списка участников перед отрисовкой
            valid_participants = [p for p in participants if p is not None]
            num_valid_participants = len(valid_participants)
            if num_valid_participants == 0: continue

            bracket_size = len(participants) # Размер сетки (степень двойки)
            num_rounds = math.ceil(math.log2(bracket_size)) if bracket_size > 1 else 0

            bracket_start_row = current_row
            max_row_drawn_this_bracket = bracket_start_row # Отслеживаем самую нижнюю строку

            # --- Заголовок категории ---
            class_name, gender, age_cat_name_key, weight_name = category_key
            category_header_row = current_row
            gender_text = "муж." if gender == "Мужской" else "жен."
            age_text = f"{age_cat_name_key} лет" if age_cat_name_key else ""
            category_title_parts = [f"Категория {class_name.strip()}" if class_name and class_name.strip() else "", gender_text, age_text]
            age_group_text = ", ".join(filter(None, category_title_parts))

            sheet.cell(current_row, 1, f"Весовая категория {weight_name} кг").font = HEADER_FONT
            # Объединяем ячейки заголовка категории
            merge_end_col = get_column_letter(max(1, 1 + num_rounds * 2)) # До куда объединять
            sheet.merge_cells(f'C{current_row}:{merge_end_col}{current_row}')
            cell = sheet.cell(current_row, 3, age_group_text)
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN
            current_row += CATEGORY_HEADER_ROWS # Переходим к началу отрисовки сетки
            max_row_drawn_this_bracket = max(max_row_drawn_this_bracket, current_row -1)


            # --- Отрисовка сетки ---
            grid_drawing_start_row = current_row
            row_coords = {} # {round_number: [y_coords_of_winners_lines]}

            if num_valid_participants == 1:
                # Сетка с одним участником
                winner = valid_participants[0]
                row1 = grid_drawing_start_row
                sheet.row_dimensions[row1].height = PARTICIPANT_ROW_HEIGHT
                sheet.row_dimensions[row1 + 1].height = PARTICIPANT_ROW_HEIGHT

                # Имя победителя сразу в последней колонке
                winner_col = 1 + num_rounds * 2
                winner_rank_title = winner.get('rank_title')
                winner_rank_suffix = _rank_suffix(winner_rank_title)
                if winner_rank_suffix:
                    winner_fio_with_rank = f"{winner.get('fio', '')} {winner_rank_suffix}"
                else:
                    winner_fio_with_rank = winner.get('fio', '')
                name_cell = sheet.cell(row1, winner_col, winner_fio_with_rank)
                name_cell.font = PARTICIPANT_FONT
                name_cell.border = BORDER_TOP_BOTTOM
                sheet.cell(row1 + 1, winner_col, f"{winner.get('city_name', '')}, {winner.get('club_name', '')}").font = SMALL_FONT
                max_row_drawn_this_bracket = max(max_row_drawn_this_bracket, row1 + 1)
            elif bracket_size >= 2:
                # Сетка с 2 и более участниками (включая BYE)
                r1_y_lines = []  # Y-координаты линий ПОД именами в первом раунде
                for i, p in enumerate(participants):
                    # Расчет Y-координаты центра блока участника
                    base_row_num = i // 2
                    is_top_in_pair = i % 2 == 0
                    row_offset = base_row_num * (
                                PARTICIPANT_ROW_HEIGHT // LINE_ROW_HEIGHT * 2 + 6)  # Увеличили отступ между парами
                    if is_top_in_pair:
                        row1 = grid_drawing_start_row + row_offset
                    else:
                        row1 = grid_drawing_start_row + row_offset + (
                                    PARTICIPANT_ROW_HEIGHT // LINE_ROW_HEIGHT + 2)  # Отступ внутри пары

                    sheet.row_dimensions[row1].height = PARTICIPANT_ROW_HEIGHT
                    sheet.row_dimensions[row1 + 1].height = PARTICIPANT_ROW_HEIGHT
                    line_y_row = row1 + (PARTICIPANT_ROW_HEIGHT // LINE_ROW_HEIGHT)  # Строка для линии под блоком

                    # --- Проверка на BYE перед отрисовкой в первой колонке ---
                    opponent_idx = i + 1 if is_top_in_pair else i - 1
                    has_opponent = 0 <= opponent_idx < len(participants)
                    opponent_is_none = has_opponent and participants[opponent_idx] is None
                    current_is_none = p is None

                    should_draw_in_col1 = True
                    if not current_is_none and opponent_is_none:  # Участник получает BYE
                        should_draw_in_col1 = False
                    elif current_is_none and has_opponent and not opponent_is_none:  # Это пустой слот напротив BYE
                        should_draw_in_col1 = False
                    # --- Конец проверки ---

                    if should_draw_in_col1:
                        if p:
                            rank_title = p.get('rank_title')
                            rank_suffix = _rank_suffix(rank_title)
                            if rank_suffix:
                                fio_with_rank = f"{p['fio']} {rank_suffix}"
                            else:
                                fio_with_rank = p['fio']
                            name_cell = sheet.cell(row1, 1, fio_with_rank)
                            name_cell.font = PARTICIPANT_FONT
                            name_cell.border = BORDER_TOP_BOTTOM
                            sheet.cell(row1 + 1, 1,
                                       f"{p.get('city_name', '')}, {p.get('club_name', '')}").font = SMALL_FONT
                            sheet.cell(row1 + 1, 1).border = Border(bottom=THIN_SIDE)
                        else:
                            name_cell = sheet.cell(row1, 1, "BYE")
                            name_cell.font = PARTICIPANT_FONT
                            name_cell.alignment = CENTER_ALIGN
                            name_cell.border = BORDER_TOP_BOTTOM

                        # Рисуем линию под блоком имени/клуба только если отрисовали блок
                        for col_idx in range(1, 2):  # Только под именем
                            sheet.cell(line_y_row, col_idx).border = Border(bottom=THIN_SIDE)

                    r1_y_lines.append(line_y_row)
                    max_row_drawn_this_bracket = max(max_row_drawn_this_bracket, line_y_row)

                # Соединительные линии и победители Раунда 1
                r2_y_coords = []
                x_start_r1 = 1 + 1  # Колонка коннектора (B)
                x_end_r1 = x_start_r1 + 1  # Колонка победителя (C)
                for i in range(0, bracket_size, 2):
                    p1 = participants[i]  # Возвращено имя participants
                    p2 = participants[i + 1]  # Возвращено имя participants
                    y1_line = r1_y_lines[i]
                    y2_line = r1_y_lines[i + 1]
                    mid_y_row = (y1_line + y2_line) // 2

                    if p1 and p2: # Стандартный бой
                        # V-линия в колонке коннектора (B)
                        for r_idx in range(y1_line, y2_line + 1):
                            sheet.cell(r_idx, x_start_r1).border = BORDER_LEFT
                        # Горизонтальная линия от середины V к победителю (C)
                        sheet.cell(mid_y_row, x_start_r1).border = BORDER_LEFT_BOTTOM # Угол + горизонталь
                        sheet.cell(mid_y_row, x_end_r1).border = BORDER_BOTTOM # Продолжение горизонтали

                        # Ячейка для имени победителя R1 (C)
                        sheet.row_dimensions[mid_y_row].height = PARTICIPANT_ROW_HEIGHT
                        sheet.cell(mid_y_row, x_end_r1).border = BORDER_TOP_BOTTOM # Рамка имени
                        sheet.row_dimensions[mid_y_row + 1].height = PARTICIPANT_ROW_HEIGHT # Для клуба/города ниже
                        sheet.cell(mid_y_row + 1, x_end_r1).border = Border(bottom=THIN_SIDE) # Нижняя рамка клуба


                        r2_y_coords.append(mid_y_row)
                    elif p1 or p2:
                        winner = p1 if p1 else p2
                        line_y_bye = y1_line if p1 else y2_line
                        mid_y_row = line_y_bye

                        rank_title = winner.get('rank_title')
                        rank_suffix = _rank_suffix(rank_title)
                        if rank_suffix:
                            fio_with_rank = f"{winner['fio']} {rank_suffix}"
                        else:
                            fio_with_rank = winner['fio']

                        name_cell = sheet.cell(mid_y_row, x_end_r1, fio_with_rank)
                        name_cell.font = PARTICIPANT_FONT
                        name_cell.border = BORDER_TOP_BOTTOM
                        sheet.row_dimensions[mid_y_row].height = PARTICIPANT_ROW_HEIGHT

                        sheet.cell(mid_y_row + 1, x_end_r1,
                                   f"{winner.get('city_name', '')}, {winner.get('club_name', '')}").font = SMALL_FONT
                        sheet.row_dimensions[mid_y_row + 1].height = PARTICIPANT_ROW_HEIGHT
                        sheet.cell(mid_y_row + 1, x_end_r1).border = Border(bottom=THIN_SIDE)

                        r2_y_coords.append(mid_y_row)
                    else:
                         r2_y_coords.append(mid_y_row)

                    max_row_drawn_this_bracket = max(max_row_drawn_this_bracket, mid_y_row + 1) # Учитываем строку клуба

                row_coords[2] = r2_y_coords # Координаты линий победителей РАУНДА 1 (для Раунда 2)

                # Раунды 2 и далее
                for r in range(2, num_rounds + 1):
                    col_connector = 1 + (r - 1) * 2 + 1 # D, F, H...
                    col_winner_box = col_connector + 1   # E, G, I...
                    prev_round_rows_coords = row_coords.get(r)
                    if not prev_round_rows_coords: break # Ошибка или конец сетки

                    current_round_rows_coords = []

                    for i in range(0, len(prev_round_rows_coords), 2):
                        y1_row = prev_round_rows_coords[i]
                        if i + 1 < len(prev_round_rows_coords): # Есть пара для боя
                            y2_row = prev_round_rows_coords[i+1]
                            mid_y_row = (y1_row + y2_row) // 2

                            # V-линия в колонке коннектора
                            for row_idx in range(y1_row, y2_row + 1):
                                sheet.cell(row_idx, col_connector).border = BORDER_LEFT
                            # Горизонтальная линия от середины V к победителю
                            sheet.cell(mid_y_row, col_connector).border = BORDER_LEFT_BOTTOM
                            sheet.cell(mid_y_row, col_winner_box).border = BORDER_BOTTOM

                            # Ячейка для имени победителя
                            sheet.row_dimensions[mid_y_row].height = PARTICIPANT_ROW_HEIGHT
                            sheet.cell(mid_y_row, col_winner_box).border = BORDER_TOP_BOTTOM
                            sheet.row_dimensions[mid_y_row + 1].height = PARTICIPANT_ROW_HEIGHT
                            sheet.cell(mid_y_row + 1, col_winner_box).border = Border(bottom=THIN_SIDE)


                            current_round_rows_coords.append(mid_y_row)
                            max_row_drawn_this_bracket = max(max_row_drawn_this_bracket, mid_y_row + 1)
                        else: # Участник без пары в этом раунде (проходит дальше)
                            # Просто рисуем рамку для имени в следующей колонке
                             sheet.row_dimensions[y1_row].height = PARTICIPANT_ROW_HEIGHT
                             sheet.cell(y1_row, col_winner_box).border = BORDER_TOP_BOTTOM
                             sheet.row_dimensions[y1_row + 1].height = PARTICIPANT_ROW_HEIGHT
                             sheet.cell(y1_row + 1, col_winner_box).border = Border(bottom=THIN_SIDE)

                             # Линия перехода в следующий этап (если он есть)
                             if r < num_rounds:
                                 next_connector_col = col_winner_box + 1
                                 sheet.cell(y1_row, col_winner_box).border = Border(top=THIN_SIDE, bottom=THIN_SIDE) # Переопределяем правую границу на пусто
                                 sheet.cell(y1_row, next_connector_col).border = BORDER_LEFT_BOTTOM # Угол + горизонталь в след. колонке

                             current_round_rows_coords.append(y1_row)
                             max_row_drawn_this_bracket = max(max_row_drawn_this_bracket, y1_row + 1)

                    if not current_round_rows_coords: break
                    row_coords[r + 1] = current_round_rows_coords


            # --- Блок подписей ---
            # Рассчитываем строку для подписей ПОСЛЕ самой нижней точки сетки
            sign_row = max_row_drawn_this_bracket + 1 + SIGNATURE_PADDING_ROWS

            # Размещаем подписи в колонках K и L (11 и 12)
            sig_col1 = 11
            sig_col2 = 12

            cell = sheet.cell(sign_row, sig_col1, "Гл. судья")
            cell.font = PARTICIPANT_FONT # Обычный шрифт
            cell.alignment = RIGHT_ALIGN
            sheet.cell(sign_row, sig_col2, "_________________Амирханов Р.Г. (г. Уфа, СВС)").font = PARTICIPANT_FONT
            max_row_drawn_this_bracket = max(max_row_drawn_this_bracket, sign_row)
            sign_row += 1

            # Удалена строка Зам гл судьи

            cell = sheet.cell(sign_row, sig_col1, "Гл. секретарь")
            cell.font = PARTICIPANT_FONT
            cell.alignment = RIGHT_ALIGN
            sheet.cell(sign_row, sig_col2, "______________ Ягафаров А. Б.(г.Стерлитамак, С1С)").font = PARTICIPANT_FONT
            max_row_drawn_this_bracket = max(max_row_drawn_this_bracket, sign_row)

            # --- Блок с местами ---
            place_row = sign_row + 4
            num_valid_participants = len(
                [p for p in participants if p is not None])

            if num_valid_participants == 1:
                places = ["1-е место"]
            elif num_valid_participants == 2:
                places = ["1-е место", "2-е место"]
            elif num_valid_participants == 3:
                places = ["1-е место", "2-е место", "3-е место"]
            else:  # 4 и более
                places = ["1 место", "2 место", "3 место", "3 место"]

            # Заголовок "Места"
            cell = sheet.cell(place_row - 1, sig_col1, "Места")
            cell.font = HEADER_FONT  # Используем жирный шрифт заголовка
            cell.alignment = RIGHT_ALIGN
            max_row_drawn_this_bracket = max(max_row_drawn_this_bracket, place_row - 1)

            # Строки с местами и линиями для подписи
            for place_text in places:
                cell = sheet.cell(place_row, sig_col1, place_text)
                cell.font = PARTICIPANT_FONT  # Обычный шрифт
                cell.alignment = RIGHT_ALIGN
                sheet.cell(place_row, sig_col2, "_____________").font = PARTICIPANT_FONT  # Линия для подписи/ФИО
                max_row_drawn_this_bracket = max(max_row_drawn_this_bracket, place_row)
                place_row += 1

            # --- Переход к следующей сетке на листе ---
            # Обновляем current_row на основе самой нижней строки (теперь это строка с местами) + отступ
            current_row = max_row_drawn_this_bracket + 1 + GRID_PADDING_ROWS

            # Сохраняем книгу ПОСЛЕ обработки ВСЕХ возрастных категорий
        try:
            wb.save(file_path)
        except Exception as e:
            print(f"Ошибка при сохранении Excel файла '{file_path}': {e}")


# Добавляем thin_border для использования в ячейках
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

def generate_pairs_list_excel(grouped_by_category: dict, file_path: str):
    """
    Генерирует Excel-файл со списком пар участников, отсортированный по стадиям (1/8, 1/4...).
    Участники без пары в первом раунде выносятся на отдельный лист.
    Применяет форматирование шрифта и выравнивания ко всем ячейкам данных.
    Исправлена сортировка по стадиям и проблема "Победитель боя None".
    Информация об участнике (ФИО и Клуб/Город) разделена на две ячейки.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Состав пар"

    unpaired_sheet = workbook.create_sheet(title="Без пары")
    unpaired_headers = ["ФИО", "Клуб", "Город", "Возрастная категория", "Вес", "Тренер"] # Добавили Город
    unpaired_sheet.append(unpaired_headers)
    for cell in unpaired_sheet[1]:
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    unpaired_for_sheet = []

    # --- Стили и шрифты ---
    bold_font_calibri_11 = Font(name='Calibri', size=11, bold=True)
    regular_font_calibri_11 = Font(name='Calibri', size=11)
    center_top_align_wrap = Alignment(horizontal='center', vertical='top', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center')
    center_bottom_align = Alignment(horizontal='center', vertical='bottom')

    # --- Цвета заливки ---
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Светло-красный
    blue_fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")  # Светло-синий


    # --- Шапка документа ---
    sheet.merge_cells('A1:F1')
    sheet['A1'] = "Чемпионат и Первенство республики Башкортостан по муайтай"
    sheet['A1'].font = regular_font_calibri_11
    sheet['A1'].alignment = center_align
    sheet.merge_cells('A2:F2')
    sheet['A2'] = "31.10 - 03.11.2025"
    sheet['A2'].font = regular_font_calibri_11
    sheet['A2'].alignment = center_align

    # --- Заголовки таблицы "Состав пар" (изменены D и E) ---
    headers = ["№ п/п", "Возрастная\nкатегория", "Весовая\nкатегория", "Красный угол (Red)",
               "Синий угол (Blue)", "Результат"]
    sheet.append(headers)
    sheet.row_dimensions[3].height = 30 # Высота строки заголовка
    for col_idx, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col_idx)
        cell.font = bold_font_calibri_11
        cell.alignment = center_top_align_wrap # Используем новое выравнивание
        cell.border = thin_border # Добавляем границы


    # --- Фаза 1: Сбор данных о парах ---
    all_pairs_data = []
    placeholder_map = {}
    pair_id_counter = 0

    sorted_categories = sorted(
        grouped_by_category.items(),
        key=lambda item: (
            _age_sort_key(item[0][2]), 0 if item[0][1] == 'Женский' else 1,
            _weight_sort_key(item[0][3]), _class_sort_key(item[0][0])
        )
    )

    for category_key, participants in sorted_categories:
        if not participants: continue
        valid_participants = [p for p in participants if p is not None]
        if len(valid_participants) == 1:
            unpaired_for_sheet.append({'p': valid_participants[0], 'key': category_key})
            continue

        seeded = participants
        bracket_size = len(seeded)
        if bracket_size < 1: continue

        num_rounds = math.ceil(math.log2(bracket_size)) if bracket_size > 1 else 0
        if num_rounds < 1:
             if len(valid_participants) == 1:
                 unpaired_for_sheet.append({'p': valid_participants[0], 'key': category_key})
             continue

        class_name, gender, age_cat_name, weight_name = category_key
        class_letter = class_name.split(' ')[0] if class_name else ""
        weight_text = f"{weight_name} {class_letter}"

        def get_participant_ref(p):
            return p if isinstance(p, dict) else None

        next_round_participants = []

        # Раунд 1
        round_num = 1
        for i in range(0, bracket_size, 2):
            p1 = seeded[i] if i < len(seeded) else None
            p2 = seeded[i + 1] if i + 1 < len(seeded) else None

            if p1 and p2:
                pair_id_counter += 1
                pair_id = f"pair_{pair_id_counter}"
                pair_info = {
                    "id": pair_id, "category_key": category_key, "age_cat_name": age_cat_name,
                    "weight_text": weight_text, "p1": get_participant_ref(p1), "p2": get_participant_ref(p2),
                    "p1_source_id": None, "p2_source_id": None, "num_rounds": num_rounds,
                    "round_num": round_num, "final_pair_num": None
                }
                next_round_participants.append(pair_id)
                all_pairs_data.append(pair_info)
                placeholder_map[pair_id] = pair_info
            elif p1:
                # Участник один в первом раунде (нечетное количество) - действительно без пары
                # Добавляем в unpaired_for_sheet только если это первый раунд
                next_round_participants.append(p1)
                if get_participant_ref(p1):
                    unpaired_for_sheet.append({'p': p1, 'key': category_key})
            elif p2:
                # Участник один в первом раунде (нечетное количество) - действительно без пары
                # Добавляем в unpaired_for_sheet только если это первый раунд
                next_round_participants.append(p2)
                if get_participant_ref(p2):
                    unpaired_for_sheet.append({'p': p2, 'key': category_key})

        # Последующие раунды
        for round_num in range(2, num_rounds + 1):
            current_round_participants = next_round_participants
            next_round_participants = []
            if len(current_round_participants) < 1: break

            for i in range(0, len(current_round_participants), 2):
                p1_ref = current_round_participants[i]
                if i + 1 < len(current_round_participants):
                    p2_ref = current_round_participants[i + 1]
                    pair_id_counter += 1
                    pair_id = f"pair_{pair_id_counter}"
                    pair_info = {
                        "id": pair_id, "category_key": category_key, "age_cat_name": age_cat_name,
                        "weight_text": weight_text, "p1": get_participant_ref(p1_ref),
                        "p2": get_participant_ref(p2_ref),
                        "p1_source_id": p1_ref if isinstance(p1_ref, str) else None,
                        "p2_source_id": p2_ref if isinstance(p2_ref, str) else None,
                        "num_rounds": num_rounds, "round_num": round_num, "final_pair_num": None
                    }
                    next_round_participants.append(pair_id)
                    all_pairs_data.append(pair_info)
                    placeholder_map[pair_id] = pair_info
                else:
                    next_round_participants.append(p1_ref)

    # --- Фаза 2: Сортировка и нумерация ---
    all_pairs_data.sort(key=lambda item: (
         _get_stage_sort_key(item['num_rounds'], item['round_num']),
         _age_sort_key(item['category_key'][2]),
         0 if item['category_key'][1] == 'Женский' else 1,
         _weight_sort_key(item['category_key'][3]),
         _class_sort_key(item['category_key'][0])
    ))
    pair_counter = 1
    for pair_info in all_pairs_data:
        pair_info["final_pair_num"] = pair_counter
        pair_counter += 1

    # --- Фаза 3: Запись в Excel ---
    current_data_row = 4 # Начинаем с 4-й строки
    for pair_info in all_pairs_data:
        row1 = current_data_row
        row2 = current_data_row + 1
        sheet.row_dimensions[row1].height = 15 # Стандартная высота для ФИО
        sheet.row_dimensions[row2].height = 15 # Стандартная высота для Клуб/Город

        # -- Объединение ячеек для №, Веса, Результата --
        sheet.merge_cells(start_row=row1, start_column=1, end_row=row2, end_column=1)  # № п/п
        # sheet.merge_cells(start_row=row1, start_column=2, end_row=row2, end_column=2) # НЕ ОБЪЕДИНЯЕМ Возр. кат + стадия
        sheet.merge_cells(start_row=row1, start_column=3, end_row=row2, end_column=3)  # Вес
        sheet.merge_cells(start_row=row1, start_column=6, end_row=row2, end_column=6)  # Результат

        # -- Запись данных для объединенных и разделенных ячеек --
        stage_text_raw = _get_stage_text(pair_info["num_rounds"], pair_info["round_num"])
        # Убираем \n из stage_text, если он там есть (теперь он в отдельной ячейке)
        stage_text = stage_text_raw.replace('\n', '')

        age_cat_display = pair_info['age_cat_name']
        if pair_info['category_key'][1] == 'Женский':
            age_cat_display += " Ж"

        sheet.cell(row=row1, column=1, value=pair_info["final_pair_num"])  # Объединенная ячейка A
        sheet.cell(row=row1, column=2, value=age_cat_display)  # Ячейка B (верхняя) - Возраст
        sheet.cell(row=row2, column=2, value=stage_text)  # Ячейка B (нижняя) - Стадия
        sheet.cell(row=row1, column=3, value=pair_info["weight_text"])  # Объединенная ячейка C
        sheet.cell(row=row1, column=6, value="")  # Объединенная ячейка F

        # Функция для получения ФИО и Клуб/Город (остается без изменений)
        def get_participant_details(p_info, source_id):
            fio_display = ""
            club_city = ""
            if p_info:
                fio_full = p_info.get('fio', '')
                fio_short = format_fio_without_patronymic(fio_full) # Фамилия Имя (с учетом исключений)
                rank_title = p_info.get('rank_title') # Получаем разряд из данных участника
                rank_suffix = _rank_suffix(rank_title) # Сокращение или исходная запись
                if rank_suffix:
                    fio_display = f"{fio_short} {rank_suffix}"
                else:
                    fio_display = fio_short
                club_city = f"{p_info.get('club_name', '')}, {p_info.get('city_name', '')}"
            elif source_id:
                source_pair = placeholder_map.get(source_id)
                if source_pair and source_pair.get("final_pair_num") is not None:
                    fio_display = f"Победитель боя {source_pair['final_pair_num']}"
                else:
                    fio_display = f"Победитель боя {source_id} (?)"
            return fio_display, club_city

        # -- Запись данных для Красного угла (столбец D) --
        fio1, club_city1 = get_participant_details(pair_info["p1"], pair_info["p1_source_id"])
        sheet.cell(row=row1, column=4, value=fio1)  # ФИО
        sheet.cell(row=row2, column=4, value=club_city1)  # Клуб, Город

        # -- Запись данных для Синего угла (столбец E) --
        fio2, club_city2 = get_participant_details(pair_info["p2"], pair_info["p2_source_id"])
        sheet.cell(row=row1, column=5, value=fio2)  # ФИО
        sheet.cell(row=row2, column=5, value=club_city2)  # Клуб, Город

        # -- Применение стилей ко всем ячейкам этой пары --
        for r in range(row1, row2 + 1):
            for c in range(1, 7):  # Проходим по всем 6 столбцам
                cell = sheet.cell(row=r, column=c)
                cell.font = regular_font_calibri_11
                cell.border = thin_border  # Применяем границы ко всем
                # Устанавливаем выравнивание
                if c in [1, 3, 6]:  # Объединенные ячейки A, C, F
                    cell.alignment = center_align  # Выравнивание по центру для объединенных
                elif r == row1 and c == 2:  # Ячейка Возраст
                    cell.alignment = center_align  # Выравнивание по центру
                elif r == row2 and c == 2:  # Ячейка Стадия
                    cell.alignment = center_align  # Выравнивание по центру
                elif r == row1 and c in [4, 5]:  # ФИО
                    cell.alignment = center_align
                elif r == row2 and c in [4, 5]:  # Клуб/Город
                    cell.alignment = center_bottom_align

        current_data_row += 2  # Переходим к следующей паре (через 2 строки)

    # --- Заполнение листа "Без пары" ---
    # Сначала собираем множество ID участников, которые участвуют хотя бы в одной паре
    participants_in_pairs = set()
    for pair_info in all_pairs_data:
        p1_ref = pair_info.get('p1')
        p2_ref = pair_info.get('p2')
        if p1_ref and isinstance(p1_ref, dict) and p1_ref.get('id'):
            participants_in_pairs.add(p1_ref.get('id'))
        if p2_ref and isinstance(p2_ref, dict) and p2_ref.get('id'):
            participants_in_pairs.add(p2_ref.get('id'))
    
    # Фильтруем unpaired_for_sheet: исключаем тех, кто участвует в парах
    truly_unpaired = []
    for item in unpaired_for_sheet:
        p = item.get('p')
        if not p: continue
        participant_id = p.get('id')
        # Если участник участвует хотя бы в одной паре, не добавляем его в "Без пары"
        if participant_id and participant_id in participants_in_pairs:
            continue
        truly_unpaired.append(item)
    
    if truly_unpaired:
        truly_unpaired.sort(key=lambda item: (
             _age_sort_key(item['key'][2]), 0 if item['key'][1] == 'Женский' else 1,
             _weight_sort_key(item['key'][3]), _class_sort_key(item['key'][0]),
             item['p'].get('fio', '') if item.get('p') else ''
        ))
        for item in truly_unpaired:
            p = item.get('p')
            category_key_unp = item.get('key')
            if not p or not category_key_unp: continue

            class_name_unp, gender_unp, age_cat_name_unp, weight_name_unp = category_key_unp
            class_letter_unp = class_name_unp.split(' ')[0] if class_name_unp else ""
            weight_text_unp = f"{weight_name_unp} {class_letter_unp}"
            age_cat_display_unp = age_cat_name_unp
            if gender_unp == 'Женский':
                age_cat_display_unp += " Ж"
            unpaired_sheet.append([
                p.get('fio', ''), p.get('club_name', 'Без клуба'), p.get('city_name', ''), # Добавили Город
                age_cat_display_unp, weight_text_unp, p.get('coach_name', '')
            ])
        # Автоподбор ширины и стилизация для листа "Без пары"
        for column_cells in unpaired_sheet.columns:
            try:
                 max_length = max(len(str(cell.value or "")) for cell in column_cells if cell.value is not None)
                 width_to_set = max(max_length + 2, 10) # Мин. ширина 10
                 column_letter = get_column_letter(column_cells[0].column)
                 unpaired_sheet.column_dimensions[column_letter].width = width_to_set
            except ValueError: pass
        for row in unpaired_sheet.iter_rows(min_row=2, max_row=unpaired_sheet.max_row):
            for cell in row:
                cell.alignment = center_align
                cell.font = regular_font_calibri_11
                cell.border = thin_border # Добавляем границы

    # Установка ширины колонок для листа "Состав пар"
    sheet.column_dimensions['A'].width = 8
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 40
    sheet.column_dimensions['E'].width = 40
    sheet.column_dimensions['F'].width = 20

    try:
        workbook.save(file_path)
    except Exception as e:
         print(f"Ошибка при сохранении Excel файла '{file_path}': {e}")

def generate_protocol_excel(participants: list, file_path: str):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Протокол"

    bold_font = Font(name='Calibri', size=11, bold=True)
    normal_font = Font(name='Calibri', size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    headers = [
        "№", "ФИО", "Дата рожд.", "Разряд",
        "Спортивная категория", "Город", "Организация", "Тренер",
        "Чел. в весе", "№ жер", "1/8", "1/4", "1/2", "Финал", "Место"
    ]
    num_headers = len(headers)
    last_header_col_letter = get_column_letter(num_headers)

    sheet.row_dimensions[1].height = 45
    sheet.append(headers)

    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for p in participants:
        age_cat = p.get('age_category_name') or "Без категории"
        gender = p.get('gender') or "Пол не указан"
        weight_val = p.get('weight')
        weight_key = _weight_sort_key(str(weight_val) if weight_val is not None else None)
        grouped_data[age_cat][gender][weight_key].append(p)

    sorted_age_cats = sorted(grouped_data.keys(), key=_age_sort_key, reverse=True)

    current_row = 2
    participant_counter = 0

    for age_cat in sorted_age_cats:
        genders_in_age = grouped_data[age_cat]
        sorted_genders = sorted(genders_in_age.keys(), reverse=True)

        for gender in sorted_genders:
            gender_label = "(муж)" if gender == "Мужской" else "(жен)"
            age_gender_header = f"{age_cat} {gender_label}"

            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_headers)
            age_gender_cell = sheet.cell(row=current_row, column=1, value=age_gender_header)
            age_gender_cell.font = bold_font
            age_gender_cell.alignment = left_align
            current_row += 1

            weights_in_gender = genders_in_age[gender]
            sorted_weight_keys = sorted(weights_in_gender.keys())

            for weight_key in sorted_weight_keys:
                participants_in_weight = weights_in_gender[weight_key]
                if not participants_in_weight: continue

                weight_display = format_weight(participants_in_weight[0].get('weight'))
                num_in_weight = len(participants_in_weight)

                merged_cell_range = f'A{current_row}:{last_header_col_letter}{current_row}'
                sheet.merge_cells(merged_cell_range)
                weight_cat_cell = sheet.cell(row=current_row, column=1, value=f"Весовая категория {weight_display}")
                weight_cat_cell.font = bold_font
                weight_cat_cell.alignment = left_align
                for row_cells in sheet[merged_cell_range]:
                    for cell in row_cells:
                        cell.fill = light_gray_fill
                current_row += 1

                for p in participants_in_weight:
                    participant_counter += 1
                    dob_formatted = p.get('dob').strftime('%d.%m.%Y') if p.get('dob') else ''

                    participant_weight_formatted = format_weight(p.get('weight')).replace(' кг', '')

                    row_data = [
                        participant_counter,
                        p.get('fio', ''),
                        dob_formatted,
                        p.get('rank_title', ''),
                        f"муайтай, {participant_weight_formatted} кг",
                        p.get('city_name', ''),
                        p.get('club_name', ''),
                        p.get('coach_name', ''),
                        num_in_weight,
                        "", "", "", "", "", ""
                    ]
                    sheet.append(row_data)
                    for col_idx in range(1, num_headers + 1):
                        cell = sheet.cell(row=current_row, column=col_idx)
                        cell.font = normal_font
                        cell.alignment = left_align
                        if col_idx in [1, 3, 4, 9] or 10 <= col_idx <= 15:
                             cell.alignment = center_align
                        if col_idx == 5:
                             cell.alignment = center_align
                        cell.border = thin_border
                    current_row += 1

    for col_idx in range(1, num_headers + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                cell_lines = str(cell_value).split('\n')
                max_length_in_cell = max(len(line) for line in cell_lines) if cell_lines else 0
                max_length = max(max_length, max_length_in_cell)

        adjusted_width = min(max(max_length + 2, 10), 40)
        sheet.column_dimensions[column_letter].width = adjusted_width
        if headers[col_idx-1] in ["ФИО", "Город", "Организация", "Тренер"]:
             sheet.column_dimensions[column_letter].width = 30
        if headers[col_idx-1] == "Спортивная категория":
             sheet.column_dimensions[column_letter].width = 25

    workbook.save(file_path)